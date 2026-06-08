"""Слой данных: каталог и выдача ключей из Google Sheets.

Структура таблицы:
  Каждый лист = игра (например, ROBLOX).
  Строка 1 — заголовки колонок вида "<количество_валюты> <цена>", например
  "100 100" (100 единиц за 100₽) или "296 200". Колонка без второго числа
  (например "450", "600") — цена не задана, такой тариф НЕ продаётся.
  Ниже заголовка в той же колонке лежат ключи — по одному в ячейке.
  Колонка "№" и листы из config.sheet_skip игнорируются.

Состояние «занято» (бронь/продано) хранится НЕ в листе, а в БД
(orders.inventory_keys) — это слой переопределения. Лист отвечает на вопрос
«какие ключи вообще есть и где», БД — «какие уже заняты». При продаже ячейка
в листе дополнительно помечается визуально (серый фон, зачёркнутый текст, ✅),
но ключ не удаляется.

Пока таблица не подключена (нет SHEET_ID / service_account.json) — каталог
пустой, бот запускается и показывает интерфейс без товаров.
"""
from __future__ import annotations

import asyncio
import datetime as dt
import logging
import re
from dataclasses import dataclass

import orders
from config import config

log = logging.getLogger("neuro_dropbot.sheets")

# Гарантирует, что бронь (read-sheet -> insert) выполняется по одному за раз
_reserve_lock = asyncio.Lock()

# Метка проданного ключа, дописываемая в ячейку (ключ при этом сохраняется)
SOLD_MARK = " ✅"
_PRODUCT_ID_SEP = "__"


@dataclass
class Product:
    id: str          # "<game>__<qty>", например "ROBLOX__100"
    title: str
    description: str
    price: int       # рублей, целое
    image_url: str = ""
    stock: int = 0
    game: str = ""
    qty: int = 0


@dataclass
class KeyCell:
    game: str
    qty: int
    price: int
    row: int         # 1-based, как в Google Sheets
    col: int         # 1-based
    key_text: str


# ---------- разбор таблицы (чистые функции, тестируются офлайн) ----------
def parse_header(header: str) -> tuple[int, int] | None:
    """'100 100' -> (qty=100, price=100). Без второго числа -> None (тариф без цены)."""
    parts = header.strip().split()
    if len(parts) < 2:
        return None
    try:
        return int(parts[0]), int(parts[1])
    except ValueError:
        return None


def clean_key(raw: str) -> str:
    """Каноничный текст ключа: без хвостовой метки продажи и пробелов."""
    val = raw.strip()
    mark = SOLD_MARK.strip()
    if mark and val.endswith(mark):
        val = val[: -len(mark)].strip()
    return val


_KEY_RE = re.compile(r"^[A-Za-z0-9][A-Za-z0-9\-]{3,}$")


def _looks_like_key(val: str) -> bool:
    return bool(_KEY_RE.match(val))


def extract_cells(game: str, values: list[list[str]]) -> list[KeyCell]:
    """Из матрицы значений листа достаёт все ключи priced-колонок."""
    if not values:
        return []
    header = values[0]
    tiers: dict[int, tuple[int, int]] = {}
    for col_idx, h in enumerate(header):
        parsed = parse_header(str(h))
        if parsed:
            tiers[col_idx] = parsed

    cells: list[KeyCell] = []
    for r in range(1, len(values)):
        row = values[r]
        for col_idx, (qty, price) in tiers.items():
            if col_idx >= len(row):
                continue
            key = clean_key(str(row[col_idx]))
            if not key or not _looks_like_key(key):
                continue
            cells.append(KeyCell(game=game, qty=qty, price=price,
                                 row=r + 1, col=col_idx + 1, key_text=key))
    return cells


def product_id(game: str, qty: int) -> str:
    return f"{game}{_PRODUCT_ID_SEP}{qty}"


def parse_product_id(pid: str) -> tuple[str, int] | None:
    if _PRODUCT_ID_SEP not in pid:
        return None
    game, qty = pid.rsplit(_PRODUCT_ID_SEP, 1)
    try:
        return game, int(qty)
    except ValueError:
        return None


# ---------- доступ к Google Sheets ----------
def _client():
    import gspread
    from google.oauth2.service_account import Credentials

    scopes = ["https://www.googleapis.com/auth/spreadsheets"]
    creds = Credentials.from_service_account_file(
        config.google_credentials_file, scopes=scopes
    )
    return gspread.authorize(creds)


def _scan_via_gspread() -> list[KeyCell]:
    """Чтение через сервисный аккаунт (когда есть креды и нужна запись)."""
    gc = _client()
    sh = gc.open_by_key(config.sheet_id)
    skip = set(config.sheet_skip)
    cells: list[KeyCell] = []
    for ws in sh.worksheets():
        if ws.title in skip:
            continue
        cells.extend(extract_cells(ws.title, ws.get_all_values()))
    return cells


def _scan_via_public() -> list[KeyCell]:
    """Чтение по публичной ссылке (xlsx-экспорт) — без сервисного аккаунта.

    Требует, чтобы таблица была открыта «Доступ по ссылке: читатель».
    """
    import io
    import requests
    import openpyxl

    r = requests.get(config.sheet_export_xlsx_url, timeout=30,
                     headers={"User-Agent": "Mozilla/5.0 (catalog-sync)"})
    r.raise_for_status()
    wb = openpyxl.load_workbook(io.BytesIO(r.content), data_only=True)
    skip = set(config.sheet_skip)
    cells: list[KeyCell] = []
    for name in wb.sheetnames:
        if name in skip:
            continue
        ws = wb[name]
        values = [
            [("" if v is None else str(v)) for v in row]
            for row in ws.iter_rows(values_only=True)
        ]
        cells.extend(extract_cells(name, values))
    return cells


# короткий кэш прочитанной таблицы, чтобы не дёргать её на каждый клик
_scan_cache: dict[str, object] = {"ts": 0.0, "cells": None}


def _scan_sheet(force: bool = False) -> list[KeyCell]:
    """Читает все игровые листы (с кэшем). Источник: креды -> gspread, иначе публичный xlsx."""
    import time

    now = time.time()
    if (not force and _scan_cache["cells"] is not None
            and now - float(_scan_cache["ts"]) < config.sheet_cache_sec):
        return _scan_cache["cells"]  # type: ignore[return-value]

    cells = _scan_via_gspread() if config.sheets_writable else _scan_via_public()
    _scan_cache["cells"] = cells
    _scan_cache["ts"] = now
    return cells


# ---------- каталог ----------
def _catalog_sync() -> list[Product]:
    if not config.sheets_readable:
        return []
    try:
        cells = _scan_sheet()
    except Exception as e:  # noqa: BLE001 — сеть/доступ к таблице не должны ронять бота
        log.warning("sheets: не удалось прочитать таблицу: %r", e)
        return []
    return _build_catalog(cells)


def _build_catalog(cells: list[KeyCell]) -> list[Product]:
    orders._inv_expire()              # снять просроченные брони
    blocked = orders._inv_blocked()   # {(game, key_text)} занятых ключей

    # агрегируем свободные по (game, qty) -> [price, count]
    agg: dict[tuple[str, int], list[int]] = {}
    for c in cells:
        if (c.game, c.key_text) in blocked:
            continue
        key = (c.game, c.qty)
        if key not in agg:
            agg[key] = [c.price, 0]
        agg[key][1] += 1

    products: list[Product] = []
    for (game, qty), (price, stock) in sorted(agg.items(), key=lambda x: (x[0][0], x[0][1])):
        products.append(Product(
            id=product_id(game, qty),
            title=f"{game} — {qty}",
            description=f"{qty} ед. игровой валюты {game}.",
            price=price,
            image_url="",
            stock=stock,
            game=game,
            qty=qty,
        ))
    return products


# ---------- бронь и продажа ----------
def _reserve_sync(game: str, qty: int, order_label: str, buyer: str) -> str | None:
    """Берёт первый свободный ключ тарифа и бронирует его в БД. None — нет свободных."""
    if not config.sheets_readable:
        return None
    try:
        cells = _scan_sheet()
    except Exception as e:  # noqa: BLE001
        log.warning("sheets: чтение таблицы при брони не удалось: %r", e)
        return None
    orders._inv_expire()
    blocked = orders._inv_blocked()
    ttl = config.reserve_ttl_min
    reserved_until = (dt.datetime.utcnow() + dt.timedelta(minutes=ttl)).strftime(
        "%Y-%m-%d %H:%M:%S"
    )
    # стабильный порядок: по строке -> детерминированно берём «верхний» ключ
    candidates = sorted(
        (c for c in cells if c.game == game and c.qty == qty), key=lambda c: c.row
    )
    for c in candidates:
        if (c.game, c.key_text) in blocked:
            continue
        ok = orders._inv_reserve(
            c.game, c.key_text, c.qty, c.price, c.row, c.col,
            buyer, order_label, reserved_until,
        )
        if ok:
            return c.key_text
        # проиграли гонку за этот ключ — пробуем следующий
        blocked.add((c.game, c.key_text))
    return None


def _mark_cell_sold(game: str, row: int, col: int, key_text: str) -> None:
    """Помечает ячейку проданной: серый фон, зачёркивание, метка ✅. Ключ сохраняется.

    Требует прав записи (сервисный аккаунт). При публичном чтении пропускается —
    статус «продано» остаётся в БД.
    """
    if not config.sheets_writable:
        return
    try:
        from gspread.utils import rowcol_to_a1

        gc = _client()
        ws = gc.open_by_key(config.sheet_id).worksheet(game)
        a1 = rowcol_to_a1(row, col)

        # сверяем, что в ячейке всё ещё нужный ключ (лист могли подвинуть)
        current = clean_key(str(ws.acell(a1).value or ""))
        if current != key_text:
            cell = ws.find(key_text)  # ищем по значению
            if cell is None:
                log.warning("sheets: ключ %s не найден в листе %s для пометки", key_text, game)
                return
            a1 = rowcol_to_a1(cell.row, cell.col)

        ws.update_acell(a1, key_text + SOLD_MARK)
        ws.format(a1, {
            "backgroundColor": {"red": 0.80, "green": 0.80, "blue": 0.80},
            "textFormat": {"strikethrough": True,
                           "foregroundColor": {"red": 0.4, "green": 0.4, "blue": 0.4}},
        })
    except Exception as e:  # noqa: BLE001 — пометка не должна ломать выдачу ключа
        log.warning("sheets: не удалось пометить ячейку проданной: %r", e)


def _confirm_sync(order_label: str, product_id_str: str, buyer: str) -> str | None:
    """Оплата подтверждена: помечаем ключ sold (БД + лист, если есть права) и возвращаем его."""
    if not config.sheets_readable:
        return None

    rec = orders._inv_confirm(order_label)
    if rec:
        _mark_cell_sold(rec["game"], rec["row"], rec["col"], rec["key_text"])
        return rec["key_text"]

    # Брони нет (истекла к моменту оплаты) — пробуем выдать свежий ключ тарифа.
    parsed = parse_product_id(product_id_str)
    if not parsed:
        return None
    game, qty = parsed
    try:
        cells = _scan_sheet()
    except Exception as e:  # noqa: BLE001
        log.warning("sheets: чтение таблицы при подтверждении не удалось: %r", e)
        return None
    orders._inv_expire()
    blocked = orders._inv_blocked()
    for c in sorted((c for c in cells if c.game == game and c.qty == qty),
                    key=lambda c: c.row):
        if (c.game, c.key_text) in blocked:
            continue
        if orders._inv_mark_sold_direct(c.game, c.key_text, c.qty, c.price,
                                        c.row, c.col, buyer, order_label):
            _mark_cell_sold(c.game, c.row, c.col, c.key_text)
            return c.key_text
        blocked.add((c.game, c.key_text))
    return None


# ---------- синхронизация цветов ячеек с БД (бронь / продажа) ----------
# Формат ячейки по статусу ключа.
_FMT_SOLD = {
    "backgroundColor": {"red": 0.80, "green": 0.80, "blue": 0.80},   # серый
    "textFormat": {"strikethrough": True,
                   "foregroundColor": {"red": 0.4, "green": 0.4, "blue": 0.4}},
}
_FMT_RESERVED = {
    "backgroundColor": {"red": 1.0, "green": 0.90, "blue": 0.40},    # янтарный — бронь
    "textFormat": {"strikethrough": False,
                   "foregroundColor": {"red": 0.0, "green": 0.0, "blue": 0.0}},
}
_FMT_FREE = {
    "backgroundColor": {"red": 1.0, "green": 1.0, "blue": 1.0},      # белый — свободно
    "textFormat": {"strikethrough": False,
                   "foregroundColor": {"red": 0.0, "green": 0.0, "blue": 0.0}},
}

# Последняя применённая «сигнатура» состояния каждого листа — чтобы не дёргать
# API, когда с прошлой синхронизации ничего не изменилось.
_marks_sig: dict[str, str] = {}
# Защита от параллельных запусков синхронизации.
_sync_running = False


def _sync_marks_sync(force: bool = False) -> None:
    """Приводит цвета ячеек листа в соответствие с БД.

    Бронь — янтарный фон, продано — серый+зачёркнуто+✅, свободно — белый.
    Идемпотентно: трогает только ячейки-ключи priced-колонок. Освободившиеся
    (истёкшая бронь / отмена) возвращаются к белому, метка ✅ снимается.
    """
    if not config.sheets_writable:
        return
    from gspread.utils import rowcol_to_a1

    orders._inv_expire()                  # снять просроченные брони
    status = orders._inv_status_map()     # {(game, key_text): 'reserved'|'sold'}
    mark = SOLD_MARK.strip()

    gc = _client()
    sh = gc.open_by_key(config.sheet_id)
    skip = set(config.sheet_skip)
    for ws in sh.worksheets():
        if ws.title in skip:
            continue
        game = ws.title
        values = ws.get_all_values()
        cells = extract_cells(game, values)
        if not cells:
            continue

        plan = [(c.row, c.col, status.get((game, c.key_text), "free"), c.key_text)
                for c in cells]
        sig = repr(sorted((r, col, st) for r, col, st, _ in plan))
        if not force and _marks_sig.get(game) == sig:
            continue

        fmt_batch: list[dict] = []
        val_batch: list[dict] = []
        for row, col, st, key_text in plan:
            a1 = rowcol_to_a1(row, col)
            raw = ""
            if row - 1 < len(values) and col - 1 < len(values[row - 1]):
                raw = str(values[row - 1][col - 1])
            has_mark = raw.strip().endswith(mark)
            if st == "sold":
                fmt_batch.append({"range": a1, "format": _FMT_SOLD})
                if not has_mark:                         # дописать ✅ к проданному
                    val_batch.append({"range": a1, "values": [[key_text + SOLD_MARK]]})
            elif st == "reserved":
                fmt_batch.append({"range": a1, "format": _FMT_RESERVED})
                if has_mark:                             # бронь временна — снять ✅
                    val_batch.append({"range": a1, "values": [[key_text]]})
            else:  # free
                fmt_batch.append({"range": a1, "format": _FMT_FREE})
                if has_mark:
                    val_batch.append({"range": a1, "values": [[key_text]]})

        if fmt_batch:
            ws.batch_format(fmt_batch)
        if val_batch:
            ws.batch_update(val_batch)
        _marks_sig[game] = sig


# ---------- async-обёртки (gspread синхронный) ----------
async def sync_marks(force: bool = False) -> None:
    """Синхронизирует цвета ячеек с БД (бронь/продажа). Без работы, если нет прав записи."""
    global _sync_running
    if not config.sheets_writable or (_sync_running and not force):
        return
    _sync_running = True
    try:
        await asyncio.to_thread(_sync_marks_sync, force)
    finally:
        _sync_running = False


def request_sync() -> None:
    """Запустить синхронизацию пометок в фоне, не блокируя текущий хэндлер."""
    if not config.sheets_writable:
        return
    try:
        asyncio.get_running_loop().create_task(sync_marks())
    except RuntimeError:
        pass  # нет работающего loop — подхватит фоновый цикл


async def get_catalog() -> list[Product]:
    return await asyncio.to_thread(_catalog_sync)


async def get_product(pid: str) -> Product | None:
    for p in await get_catalog():
        if p.id == pid:
            return p
    return None


async def reserve_for_order(pid: str, order_label: str, buyer: str) -> str | None:
    """Бронирует ключ под заказ при нажатии «Купить». None — товар закончился."""
    parsed = parse_product_id(pid)
    if not parsed:
        return None
    game, qty = parsed
    async with _reserve_lock:
        return await asyncio.to_thread(_reserve_sync, game, qty, order_label, buyer)


async def confirm_sale(order_label: str, pid: str, buyer: str) -> str | None:
    """Подтверждает продажу после оплаты: ключ -> sold, ячейка помечается."""
    async with _reserve_lock:
        return await asyncio.to_thread(_confirm_sync, order_label, pid, buyer)


async def release_order(order_label: str) -> None:
    """Снимает бронь (если заказ отменён/не оплачен принудительно)."""
    await orders.inv_release(order_label)
